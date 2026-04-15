import streamlit as st
import pandas as pd
import openpyxl
import re
import requests
import os
import zipfile
from io import BytesIO

# ───────────────────────── FILE PATHS ─────────────────────────
DEFAULT_TEMPLATE = "sku-template (4).xlsx"
FALLBACK_UPLOADED_TEMPLATE = "/mnt/data/output_template (62).xlsx"

if os.path.exists(FALLBACK_UPLOADED_TEMPLATE):
    TEMPLATE_PATH = FALLBACK_UPLOADED_TEMPLATE
else:
    TEMPLATE_PATH = DEFAULT_TEMPLATE

APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbxiCe1IVsghaaFa4zJvA-YuCowvvT3JzLZag1IAp9B8MFGk6w8hI4aBpoB_WsqWkbbLPg/exec"

try:
    if not APPS_SCRIPT_URL:
        APPS_SCRIPT_URL = st.secrets["APPS_SCRIPT_URL"]
except Exception:
    pass

_FALLBACK_FILE = "batch_id_counter.json"

def _local_read() -> int:
    import json
    if os.path.exists(_FALLBACK_FILE):
        try:
            with open(_FALLBACK_FILE) as f:
                return int(json.load(f).get("v", 1))
        except Exception:
            pass
    return 1

def _local_write(v: int):
    import json
    tmp = _FALLBACK_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump({"v": v}, f)
    os.replace(tmp, _FALLBACK_FILE)

def _remote_read() -> int:
    r = requests.get(APPS_SCRIPT_URL, timeout=10)
    r.raise_for_status()
    return int(r.json()["batch_id"])

def _remote_write(next_id: int):
    r = requests.post(APPS_SCRIPT_URL, json={"next_id": next_id}, timeout=10)
    r.raise_for_status()

def peek_next_batch_id() -> int:
    if APPS_SCRIPT_URL:
        try:
            return _remote_read()
        except Exception:
            pass
    return _local_read()

def get_and_increment_batch_id() -> int:
    if APPS_SCRIPT_URL:
        try:
            current = _remote_read()
            _remote_write(current + 1)
            return current
        except Exception:
            pass
    current = _local_read()
    _local_write(current + 1)
    return current


# ╭───────────────── NORMALISERS & HELPERS ─────────────────╮
def norm(s) -> str:
    if pd.isna(s):
        return ""
    return "".join(str(s).split()).lower()

def clean_header(header) -> str:
    if pd.isna(header):
        return ""
    header_str = str(header)
    header_str = re.sub(r"[^0-9A-Za-z ]+", " ", header_str)
    header_str = re.sub(r"\s+", " ", header_str).strip()
    return header_str

IMAGE_EXT_RE = re.compile(r"(?i)\.(jpe?g|png|gif|bmp|webp|tiff?)$")
IMAGE_KEYWORDS = {"image", "img", "picture", "photo", "thumbnail", "thumb", "hero", "front", "back", "url"}

def is_image_column(col_header_norm: str, series: pd.Series) -> bool:
    header_hit = any(k in col_header_norm for k in IMAGE_KEYWORDS)
    sample = series.dropna().astype(str).head(20)
    ratio = sample.str.contains(IMAGE_EXT_RE).mean() if not sample.empty else 0.0
    return header_hit or ratio >= 0.30

def dedupe_columns(columns):
    seen = {}
    result = []
    for col in columns:
        col_str = str(col) if not pd.isna(col) else "Unnamed"
        if col_str in seen:
            seen[col_str] += 1
            result.append(f"{col_str}_{seen[col_str]}")
        else:
            seen[col_str] = 0
            result.append(col_str)
    return result
# ╰───────────────────────────────────────────────────────────╯

MARKETPLACE_ID_MAP = {
    "Amazon":   ("Parent SKU", "SKU"),
    "Myntra":   ("styleId", "styleGroupId"),
    "Ajio":     ("*Item SKU", "*Style Code"),
    "Flipkart": ("Seller SKU ID", "Style Code"),
    "TataCliq": ("Seller Article SKU", "*Style Code"),
    "Zivame":   ("Style Code", "SKU Code"),
    "Celio":    ("Style Code", "SKU Code"),
    "Meesho":   ("Product ID / Style ID", "SKU ID"),
}

STYLE_GROUP_MAPPING = {
    "Amazon": {
        "parent": "Parent SKU",
        "color":  "Color",
        "price":  "Maximum Retail Price (Sell on Amazon, IN)",
        "image":  "Main Image URL"
    },
    "Flipkart": {
        "parent": "Style Code",
        "color":  "Brand Color",
        "price":  "MRP",
        "image":  "Main Image URL"
    },
    "Meesho": {
        "parent": "Product ID / Style ID",
        "color":  "Color",
        "price":  "MRP",
        "image":  "Image 1"
    }
}

def generate_style_group_id(df, marketplace):
    mapping = STYLE_GROUP_MAPPING.get(marketplace)
    if not mapping:
        df["styleGroupId"] = ""
        return df
    def find_exact_or_fuzzy(df, name):
        nname = norm(name)
        for c in df.columns:
            if str(c).strip() == name:
                return c
        for c in df.columns:
            if norm(str(c)) == nname:
                return c
        for c in df.columns:
            if norm(str(c)).startswith(nname):
                return c
        return None

    parent_col = find_exact_or_fuzzy(df, mapping["parent"])
    color_col  = find_exact_or_fuzzy(df, mapping["color"])
    price_col  = find_exact_or_fuzzy(df, mapping["price"])
    image_col  = find_exact_or_fuzzy(df, mapping["image"])
    if not color_col or not price_col or not image_col:
        df["styleGroupId"] = [str(i + 1) for i in range(len(df))]
        return df

    if not parent_col:
        df[color_col] = df[color_col].astype(str).str.strip().str.lower().str.replace(" ", "", regex=False)
        df[price_col] = pd.to_numeric(df[price_col], errors="coerce").fillna(0).astype(int).astype(str)
        df[image_col] = df[image_col].astype(str).str.strip()
        def _make_key_no_parent(row):
            image = row[image_col]
            color = row[color_col]
            price = row[price_col]
            if image not in ("", "nan", "None"):
                return f"{image}_{color}_{price}"
            return None
        df["_style_key"] = df.apply(_make_key_no_parent, axis=1)
        valid_keys = df["_style_key"].dropna().unique()
        key_map = {k: i + 1 for i, k in enumerate(valid_keys)}
        df["styleGroupId"] = df["_style_key"].map(key_map).fillna("").astype(str)
        df.drop(columns=["_style_key"], inplace=True)
        return df

    df[parent_col] = df[parent_col].astype(str).str.strip()
    df[color_col]  = (
        df[color_col].astype(str).str.strip().str.lower()
        .str.replace(" ", "", regex=False)
    )
    df[price_col]  = (
        pd.to_numeric(df[price_col], errors="coerce")
        .fillna(0).astype(int).astype(str)
    )
    df[image_col]  = df[image_col].astype(str).str.strip()
    _NULL_PARENT_VALS = {"", "none", "nan", "null", "n/a"}
    parent_counts  = df[parent_col][~df[parent_col].str.lower().isin(_NULL_PARENT_VALS)].value_counts()

    if not any(v > 1 for v in parent_counts.values):
        df["styleGroupId"] = [str(i + 1) for i in range(len(df))]
        return df

    parent_count_map = parent_counts.to_dict()
    _NULL_PARENTS = {"", "none", "nan", "null", "n/a"}

    def _make_key(row):
        parent = row[parent_col]
        color  = row[color_col]
        price  = row[price_col]
        image  = row[image_col]
        parent_is_valid = str(parent).strip().lower() not in _NULL_PARENTS
        if parent_is_valid and parent_count_map.get(parent, 0) > 1:
            return f"{parent}_{color}_{price}"
        elif image not in ("", "nan", "None"):
            return f"{image}_{color}_{price}"
        return None

    df["_style_key"] = df.apply(_make_key, axis=1)
    valid_keys = df["_style_key"].dropna().unique()
    key_map = {k: i + 1 for i, k in enumerate(valid_keys)}
    df["styleGroupId"] = df["_style_key"].map(key_map).fillna("").astype(str)
    df.drop(columns=["_style_key"], inplace=True)
    return df

def find_column_by_name_like(src_df: pd.DataFrame, name: str):
    if not name:
        return None
    name = str(name).strip()
    for c in src_df.columns:
        if str(c).strip() == name:
            return c
    nname = norm(name)
    for c in src_df.columns:
        if norm(c) == nname:
            return c
    for c in src_df.columns:
        if nname in norm(c):
            return c
    return None

def read_input_to_df(input_file, marketplace, header_row=1, data_row=2, sheet_name=None):
    marketplace_configs = {
        "Amazon":   {"sheet": "Template", "header_row": 4, "data_row": 7,  "sheet_index": None},
        "Flipkart": {"sheet": None,       "header_row": 1, "data_row": 5,  "sheet_index": 2},
        "Myntra":   {"sheet": None,       "header_row": 3, "data_row": 4,  "sheet_index": 1},
        "Ajio":     {"sheet": None,       "header_row": 2, "data_row": 3,  "sheet_index": 2},
        "TataCliq": {"sheet": None,       "header_row": 4, "data_row": 6,  "sheet_index": 0},
        "Meesho":   {"sheet": None,       "header_row": 3, "data_row": 6,  "sheet_index": 1},
        "General":  {"sheet": None,       "header_row": header_row, "data_row": data_row, "sheet_index": 0}
    }
    config = marketplace_configs.get(marketplace, marketplace_configs["General"])

    if marketplace == "General" and sheet_name:
        xl = pd.ExcelFile(input_file)
        temp_df = xl.parse(sheet_name, header=None)
        header_idx = header_row - 1
        data_idx = data_row - 1
        headers = temp_df.iloc[header_idx].tolist()
        src_df = temp_df.iloc[data_idx:].copy()
        src_df.columns = dedupe_columns(headers)
        src_df.reset_index(drop=True, inplace=True)

    elif config["sheet"] is not None:
        _wb = openpyxl.load_workbook(input_file, data_only=True)
        _ws = _wb[config["sheet"]]
        header_idx = config["header_row"]
        data_idx   = config["data_row"]
        headers = [_ws.cell(row=header_idx, column=c).value for c in range(1, _ws.max_column + 1)]
        data_rows_raw = []
        for r in range(data_idx, _ws.max_row + 1):
            data_rows_raw.append([_ws.cell(row=r, column=c).value for c in range(1, _ws.max_column + 1)])
        src_df = pd.DataFrame(data_rows_raw, columns=dedupe_columns(headers))
        src_df.reset_index(drop=True, inplace=True)
        if marketplace == "Amazon":
            parentage_col = find_column_by_name_like(src_df, "Parentage Level")
            if parentage_col:
                before = len(src_df)
                src_df = src_df[
                    src_df[parentage_col].astype(str).str.strip().str.lower() != "parent"
                ].copy()
                src_df.reset_index(drop=True, inplace=True)
                after = len(src_df)
                src_df.attrs["filtered_parent_rows"] = before - after
    else:
        if marketplace == "Meesho":
            import openpyxl as _oxl
            _wb = _oxl.load_workbook(input_file if not hasattr(input_file, "read") else input_file, data_only=True)
            _ws = _wb.worksheets[config["sheet_index"]]
            header_idx = config["header_row"]
            data_idx   = config["data_row"]
            headers = [_ws.cell(row=header_idx, column=c).value for c in range(1, _ws.max_column + 1)]
            data_rows = []
            for r in range(data_idx, _ws.max_row + 1):
                row_vals = [_ws.cell(row=r, column=c).value for c in range(1, _ws.max_column + 1)]
                data_rows.append(row_vals)
            src_df = pd.DataFrame(data_rows, columns=dedupe_columns(headers))
            src_df.reset_index(drop=True, inplace=True)
        else:
            xl = pd.ExcelFile(input_file)
            temp_df = xl.parse(xl.sheet_names[config["sheet_index"]], header=None)
            header_idx = config["header_row"] - 1
            data_idx = config["data_row"] - 1
            headers = temp_df.iloc[header_idx].tolist()
            src_df = temp_df.iloc[data_idx:].copy()
            src_df.columns = dedupe_columns(headers)
            src_df.reset_index(drop=True, inplace=True)

    src_df.dropna(axis=1, how='all', inplace=True)
    return src_df


FLIPKART_JOIN_COL = "Flipkart Serial Number"

def merge_flipkart_files(catalog_file=None, listing_file=None):
    def read_sheet(f, sheet_index, header_row, data_row):
        xl = pd.ExcelFile(f)
        temp = xl.parse(xl.sheet_names[sheet_index], header=None)
        headers = temp.iloc[header_row - 1].tolist()
        df = temp.iloc[data_row - 1:].copy()
        df.columns = dedupe_columns(headers)
        df.reset_index(drop=True, inplace=True)
        df.dropna(axis=1, how="all", inplace=True)
        return df

    cat_df = None
    lst_df = None

    if catalog_file is not None:
        cat_df = read_sheet(catalog_file, sheet_index=2, header_row=1, data_row=5)
    if listing_file is not None:
        lst_df = read_sheet(listing_file, sheet_index=0, header_row=1, data_row=3)

    if cat_df is None:
        return lst_df
    if lst_df is None:
        return cat_df

    def find_serial_col(df):
        col = find_column_by_name_like(df, FLIPKART_JOIN_COL)
        if col:
            return col
        for c in df.columns:
            if "serial" in norm(str(c)):
                return c
        return None

    cat_join = find_serial_col(cat_df)
    lst_join = find_serial_col(lst_df)

    if not cat_join or not lst_join:
        return cat_df

    cat_df[cat_join] = cat_df[cat_join].astype(str).str.strip()
    lst_df[lst_join] = lst_df[lst_join].astype(str).str.strip()

    cat_norms = set(norm(str(c)) for c in cat_df.columns)
    lst_cols_to_keep = [lst_join] + [
        c for c in lst_df.columns
        if c != lst_join and norm(str(c)) not in cat_norms
    ]
    lst_df_trimmed = lst_df[lst_cols_to_keep]

    merged = pd.merge(
        cat_df, lst_df_trimmed,
        left_on=cat_join, right_on=lst_join,
        how="left"
    )
    if cat_join != lst_join and lst_join in merged.columns:
        merged.drop(columns=[lst_join], inplace=True)

    merged.reset_index(drop=True, inplace=True)
    return merged

def process_file(
    input_file,
    marketplace: str,
    selected_variant_col=None,
    selected_product_col=None,
    general_header_row: int = 1,
    general_data_row: int = 2,
    general_sheet_name=None,
    premerged_df=None,
):
    if premerged_df is not None:
        src_df = premerged_df.copy()
    else:
        src_df = read_input_to_df(
            input_file, marketplace,
            header_row=general_header_row,
            data_row=general_data_row,
            sheet_name=general_sheet_name
        )
    src_df = generate_style_group_id(src_df, marketplace)

    batch_id     = get_and_increment_batch_id()
    batch_id_str = str(batch_id)
    num_rows     = len(src_df)

    _SKIP_COLS = {"styleGroupId"}
    columns_meta = []
    for col in src_df.columns:
        if str(col) in _SKIP_COLS:
            continue
        if marketplace == "Meesho":
            raw = str(col)
            parts = [p.strip() for p in raw.split("\n") if p.strip()]
            display_col = parts[0] if parts else raw.strip()
        else:
            display_col = col
        dtype = "imageurlarray" if is_image_column(norm(col), src_df[col]) else "string"
        out_col = col
        if marketplace == "Flipkart" and str(col).strip() == "Brand":
            out_col = "Brand Name"
        elif marketplace == "Meesho" and norm(col).startswith("brandname"):
            out_col = "Brand Name"
        final_out = out_col if out_col != col else display_col
        columns_meta.append({"src": col, "out": final_out, "row3": "mandatory", "row4": dtype})

    color_cols = [col for col in src_df.columns if "color" in norm(col) or "colour" in norm(col)]
    size_cols  = [col for col in src_df.columns if "size" in norm(col)]

    option1_data = pd.Series([""] * num_rows, dtype=str)
    option2_data = pd.Series([""] * num_rows, dtype=str)
    if size_cols:
        option1_data = src_df[size_cols[0]].fillna('').astype(str).str.strip()
        if color_cols and color_cols[0] != size_cols[0]:
            option2_data = src_df[color_cols[0]].fillna('').astype(str).str.strip()
    elif color_cols:
        option2_data = src_df[color_cols[0]].fillna('').astype(str).str.strip()

    unique_opt1 = option1_data.replace("", pd.NA).dropna().unique().tolist()
    unique_opt2 = option2_data.replace("", pd.NA).dropna().unique().tolist()

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_vals  = wb["Values"]
    ws_types = wb["Types"]

    def first_empty_col(ws, header_rows=(1,)):
        for col_idx in range(1, 201):
            empty = True
            for r in header_rows:
                if ws.cell(row=r, column=col_idx).value not in (None, ""):
                    empty = False
                    break
            if empty:
                return col_idx
        return ws.max_column + 1

    vals_start_col  = first_empty_col(ws_vals,  header_rows=(1,))
    types_start_col = first_empty_col(ws_types, header_rows=(1, 2, 3, 4))

    for idx, meta in enumerate(columns_meta):
        vcol = vals_start_col  + idx
        tcol = types_start_col + idx
        header_display = clean_header(meta["out"])
        ws_vals.cell(row=1, column=vcol, value=header_display)
        for r_idx, value in enumerate(src_df[meta["src"]].tolist(), start=2):
            cell = ws_vals.cell(row=r_idx, column=vcol)
            if pd.isna(value):
                cell.value = None
            else:
                if str(meta["row4"]).lower() in ("string", "imageurlarray"):
                    cell.value = str(value)
                    cell.number_format = "@"
                else:
                    cell.value = value
        ws_types.cell(row=1, column=tcol, value=header_display)
        ws_types.cell(row=2, column=tcol, value=header_display)
        ws_types.cell(row=3, column=tcol, value=meta["row3"])
        ws_types.cell(row=4, column=tcol, value=meta["row4"])

    opt1_vcol = vals_start_col  + len(columns_meta)
    opt2_vcol = opt1_vcol + 1
    ws_vals.cell(row=1, column=opt1_vcol, value="Option 1")
    ws_vals.cell(row=1, column=opt2_vcol, value="Option 2")
    for i, v in enumerate(option1_data.tolist(), start=2):
        ws_vals.cell(row=i, column=opt1_vcol, value=v if v else None)
    for i, v in enumerate(option2_data.tolist(), start=2):
        ws_vals.cell(row=i, column=opt2_vcol, value=v if v else None)

    opt1_tcol = types_start_col + len(columns_meta)
    opt2_tcol = opt1_tcol + 1
    ws_types.cell(row=1, column=opt1_tcol, value="Option 1")
    ws_types.cell(row=2, column=opt1_tcol, value="Option 1")
    ws_types.cell(row=3, column=opt1_tcol, value="non mandatory")
    ws_types.cell(row=4, column=opt1_tcol, value="select")
    ws_types.cell(row=1, column=opt2_tcol, value="Option 2")
    ws_types.cell(row=2, column=opt2_tcol, value="Option 2")
    ws_types.cell(row=3, column=opt2_tcol, value="non mandatory")
    ws_types.cell(row=4, column=opt2_tcol, value="select")
    for i, val in enumerate(unique_opt1, start=5):
        ws_types.cell(row=i, column=opt1_tcol, value=val)
    for i, val in enumerate(unique_opt2, start=5):
        ws_types.cell(row=i, column=opt2_tcol, value=val)

    def append_id_columns(variant_series, product_series):
        has_var  = variant_series is not None and variant_series.replace("", pd.NA).dropna().shape[0] > 0
        has_prod = product_series is not None and product_series.replace("", pd.NA).dropna().shape[0] > 0
        if not (has_var or has_prod):
            return
        after_written_vals  = vals_start_col  + len(columns_meta) + 2
        after_written_types = types_start_col + len(columns_meta) + 2
        cur_v = after_written_vals
        cur_t = after_written_types
        if has_var:
            ws_vals.cell(row=1, column=cur_v, value="variantId")
            for i, v in enumerate(variant_series.tolist(), start=2):
                cell = ws_vals.cell(row=i, column=cur_v, value=v if v else None)
                cell.number_format = "@"
            ws_types.cell(row=1, column=cur_t, value="variantId")
            ws_types.cell(row=2, column=cur_t, value="variantId")
            ws_types.cell(row=3, column=cur_t, value="mandatory")
            ws_types.cell(row=4, column=cur_t, value="string")
            cur_v += 1
            cur_t += 1
        if has_prod:
            ws_vals.cell(row=1, column=cur_v, value="productId")
            for i, v in enumerate(product_series.tolist(), start=2):
                cell = ws_vals.cell(row=i, column=cur_v, value=v if v else None)
                cell.number_format = "@"
            ws_types.cell(row=1, column=cur_t, value="productId")
            ws_types.cell(row=2, column=cur_t, value="productId")
            ws_types.cell(row=3, column=cur_t, value="mandatory")
            ws_types.cell(row=4, column=cur_t, value="string")

    if marketplace == "General":
        variant_series = None
        product_series = None
        if selected_variant_col and selected_variant_col != "(none)":
            if selected_variant_col in src_df.columns:
                variant_series = src_df[selected_variant_col].fillna("").astype(str)
        if selected_product_col and selected_product_col != "(none)":
            if selected_product_col in src_df.columns:
                product_series = src_df[selected_product_col].fillna("").astype(str)
        append_id_columns(variant_series, product_series)
    else:
        mapping = MARKETPLACE_ID_MAP.get(marketplace, None)
        if mapping:
            prod_src_name, var_src_name = mapping
            prod_col = find_column_by_name_like(src_df, prod_src_name)
            var_col  = find_column_by_name_like(src_df, var_src_name)
            prod_series = src_df[prod_col].fillna("").astype(str) if prod_col else None
            var_series  = src_df[var_col].fillna("").astype(str)  if var_col  else None
            append_id_columns(var_series, prod_series)

    sgi_series = src_df["styleGroupId"].astype(str).str.strip() if "styleGroupId" in src_df.columns else None
    if sgi_series is not None and sgi_series.replace("", pd.NA).dropna().shape[0] > 0:
        sgi_vcol = first_empty_col(ws_vals, header_rows=(1,))
        ws_vals.cell(row=1, column=sgi_vcol, value="styleGroupId")
        for i, v in enumerate(sgi_series.tolist(), start=2):
            cell = ws_vals.cell(row=i, column=sgi_vcol, value=v if v else None)
            cell.number_format = "@"
        sgi_tcol = first_empty_col(ws_types, header_rows=(1, 2, 3, 4))
        ws_types.cell(row=1, column=sgi_tcol, value="styleGroupId")
        ws_types.cell(row=2, column=sgi_tcol, value="styleGroupId")
        ws_types.cell(row=3, column=sgi_tcol, value="non mandatory")
        ws_types.cell(row=4, column=sgi_tcol, value="string")

    has_data_mask = src_df.apply(
        lambda row: any(
            str(v).strip() not in ("", "nan", "None")
            for v in row
            if v is not None and not (isinstance(v, float) and __import__("math").isnan(v))
        ),
        axis=1
    )
    bv_col = first_empty_col(ws_vals, header_rows=(1,))
    ws_vals.cell(row=1, column=bv_col, value="BatchID")
    for df_idx, has_data in enumerate(has_data_mask):
        if has_data:
            r = df_idx + 2
            cell = ws_vals.cell(row=r, column=bv_col, value=batch_id_str)
            cell.number_format = "@"

    bt_col = first_empty_col(ws_types, header_rows=(1, 2, 3, 4))
    ws_types.cell(row=1, column=bt_col, value="BatchID")
    ws_types.cell(row=2, column=bt_col, value="BatchID")
    ws_types.cell(row=3, column=bt_col, value="non mandatory")
    ws_types.cell(row=4, column=bt_col, value="string")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, batch_id


def get_output_filename(input_name: str, batch_id: int) -> str:
    """Generate output filename: Output_<original_name_without_ext>_batch<id>.xlsx"""
    base = os.path.splitext(input_name)[0]
    return f"Output_{base}_batch{batch_id}.xlsx"


def build_zip(file_buffers: list) -> BytesIO:
    """Pack list of (filename, BytesIO) into a ZIP buffer."""
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, buf in file_buffers:
            buf.seek(0)
            zf.writestr(fname, buf.read())
    zip_buf.seek(0)
    return zip_buf


# ───────────────────────── STREAMLIT UI ─────────────────────────
st.set_page_config(page_title="SKU Template Automation", layout="wide")
st.title("Rubick OS Template Conversion")

if not APPS_SCRIPT_URL:
    st.warning(
        "⚠️ **BatchID Google Sheets sync not configured.**  \n"
        "Set `APPS_SCRIPT_URL` in the code (or `secrets.toml`) after completing the "
        "one-time Apps Script setup described in the code comments.  \n"
        "Until then, BatchID falls back to a local file counter."
    )

next_id = peek_next_batch_id()
st.info(f"📦 Next BatchID to be assigned: **{next_id}**")

if os.path.exists(TEMPLATE_PATH):
    st.info(f"Using template: {os.path.basename(TEMPLATE_PATH)}")
    try:
        with open(TEMPLATE_PATH, "rb") as f:
            st.download_button(
                "Download current template (for reference)",
                data=f.read(),
                file_name=os.path.basename(TEMPLATE_PATH),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception:
        pass

marketplace_options = ["General", "Amazon", "Flipkart", "Myntra", "Ajio", "TataCliq", "Zivame", "Celio", "Meesho"]
marketplace_type = st.selectbox("Select Template Type", marketplace_options)

# ── Bulk mode toggle ───────────────────────────────────────────
bulk_mode = st.toggle(
    "📦 Bulk Upload Mode",
    value=False,
    help="Upload multiple files at once. All outputs will be bundled into a single ZIP download."
)

general_header_row = 1
general_data_row   = 2

if marketplace_type == "General":
    st.info("Callout: Leave blank to use defaults — Header row = 1, Data row = 2.")
    col_h, col_d = st.columns(2)
    with col_h:
        _hr = st.text_input("Header row (1-indexed)", value="", placeholder="Default: 1")
    with col_d:
        _dr = st.text_input("Data row (1-indexed)", value="", placeholder="Default: 2")
    try:
        general_header_row = int(_hr.strip()) if _hr.strip() else 1
    except ValueError:
        st.error("Header row must be a number.")
        general_header_row = 1
    try:
        general_data_row = int(_dr.strip()) if _dr.strip() else 2
    except ValueError:
        st.error("Data row must be a number.")
        general_data_row = 2

# ══════════════════════════════════════════════════════════════
#  BULK MODE
# ══════════════════════════════════════════════════════════════
if bulk_mode:
    st.markdown("---")
    st.subheader("📦 Bulk Upload")

    if marketplace_type == "Flipkart":
        st.info(
            "For Flipkart bulk mode: upload all **Catalog** files together and all **Listing** files together. "
            "Files are paired by matching name prefix (e.g. `brand_catalog.xlsx` ↔ `brand_listing.xlsx`). "
            "Unmatched files are processed solo."
        )
        col_cat, col_lst = st.columns(2)
        with col_cat:
            bulk_catalog_files = st.file_uploader(
                "Catalog Files", type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True, key="bulk_fk_catalog"
            )
        with col_lst:
            bulk_listing_files = st.file_uploader(
                "Listing Files", type=["xlsx", "xls", "xlsm"],
                accept_multiple_files=True, key="bulk_fk_listing"
            )
        all_bulk_files = bulk_catalog_files or []
        has_files = bool(bulk_catalog_files or bulk_listing_files)
    else:
        bulk_files = st.file_uploader(
            f"Upload Multiple Files ({marketplace_type})",
            type=["xlsx", "xls", "xlsm"],
            accept_multiple_files=True,
            key="bulk_files"
        )
        has_files = bool(bulk_files)

    if has_files:
        st.markdown(f"**Files ready to process:**")

        if marketplace_type == "Flipkart":
            # Build pairs: match catalog ↔ listing by stem similarity
            def stem(f):
                return os.path.splitext(f.name)[0].lower()

            listing_map = {stem(f): f for f in (bulk_listing_files or [])}
            pairs = []
            unmatched_catalogs = []
            for cf in (bulk_catalog_files or []):
                cf_stem = stem(cf)
                # Try to find a listing whose stem shares most characters
                matched = None
                for ls, lf in listing_map.items():
                    # simple check: one stem is a substring of the other, or they share >60% chars
                    if ls in cf_stem or cf_stem in ls:
                        matched = lf
                        break
                if matched:
                    pairs.append((cf, matched))
                    listing_map.pop(stem(matched), None)
                else:
                    unmatched_catalogs.append((cf, None))

            # Remaining unmatched listings
            unmatched_listings = [(None, lf) for lf in listing_map.values()]
            all_pairs = pairs + unmatched_catalogs + unmatched_listings

            for cf, lf in all_pairs:
                cat_name = cf.name if cf else "—"
                lst_name = lf.name if lf else "—"
                st.write(f"• Catalog: `{cat_name}` ↔ Listing: `{lst_name}`")

            if st.button("🚀 Process All & Download ZIP", key="bulk_fk_go"):
                results = []
                errors = []
                progress = st.progress(0)
                status_text = st.empty()

                for i, (cf, lf) in enumerate(all_pairs):
                    label = (cf or lf).name
                    status_text.text(f"Processing {i+1}/{len(all_pairs)}: {label}…")
                    try:
                        merged = merge_flipkart_files(catalog_file=cf, listing_file=lf)
                        buf, bid = process_file(
                            cf or lf, "Flipkart",
                            premerged_df=merged
                        )
                        out_name = get_output_filename(label, bid)
                        results.append((out_name, buf))
                    except Exception as e:
                        errors.append(f"`{label}`: {e}")
                    progress.progress((i + 1) / len(all_pairs))

                status_text.empty()
                progress.empty()

                if results:
                    first_bid = int(results[0][0].split("batch")[1].replace(".xlsx", ""))
                    last_bid  = int(results[-1][0].split("batch")[1].replace(".xlsx", ""))
                    zip_buf = build_zip(results)
                    st.success(f"✅ {len(results)} file(s) processed! BatchIDs {first_bid}–{last_bid}")
                    st.download_button(
                        f"📥 Download ZIP ({len(results)} files)",
                        data=zip_buf,
                        file_name=f"bulk_outputs_batch{first_bid}_to_{last_bid}.zip",
                        mime="application/zip",
                        key="bulk_fk_download"
                    )
                if errors:
                    st.error("Some files failed:")
                    for e in errors:
                        st.markdown(f"- {e}")

        elif marketplace_type == "General":
            # For General bulk: use same header/data row settings for all files
            # Show file list
            for f in bulk_files:
                st.write(f"• `{f.name}`")

            st.info(
                "💡 In bulk General mode, all files use the same header row, data row, "
                "and the **first sheet** by default. variantId/productId column selection is skipped."
            )

            if st.button("🚀 Process All & Download ZIP", key="bulk_gen_go"):
                results = []
                errors = []
                progress = st.progress(0)
                status_text = st.empty()

                for i, f in enumerate(bulk_files):
                    status_text.text(f"Processing {i+1}/{len(bulk_files)}: {f.name}…")
                    try:
                        # Use first sheet for bulk general mode
                        xl = pd.ExcelFile(f)
                        first_sheet = xl.sheet_names[0]
                        buf, bid = process_file(
                            f, "General",
                            general_header_row=general_header_row,
                            general_data_row=general_data_row,
                            general_sheet_name=first_sheet,
                        )
                        out_name = get_output_filename(f.name, bid)
                        results.append((out_name, buf))
                    except Exception as e:
                        errors.append(f"`{f.name}`: {e}")
                    progress.progress((i + 1) / len(bulk_files))

                status_text.empty()
                progress.empty()

                if results:
                    first_bid = int(results[0][0].split("batch")[1].replace(".xlsx", ""))
                    last_bid  = int(results[-1][0].split("batch")[1].replace(".xlsx", ""))
                    zip_buf = build_zip(results)
                    st.success(f"✅ {len(results)} file(s) processed! BatchIDs {first_bid}–{last_bid}")
                    st.download_button(
                        f"📥 Download ZIP ({len(results)} files)",
                        data=zip_buf,
                        file_name=f"bulk_outputs_batch{first_bid}_to_{last_bid}.zip",
                        mime="application/zip",
                        key="bulk_gen_download"
                    )
                if errors:
                    st.error("Some files failed:")
                    for e in errors:
                        st.markdown(f"- {e}")

        else:
            # All other marketplaces: straightforward multi-file upload
            for f in bulk_files:
                st.write(f"• `{f.name}`")

            if st.button("🚀 Process All & Download ZIP", key="bulk_go"):
                results = []
                errors = []
                progress = st.progress(0)
                status_text = st.empty()

                for i, f in enumerate(bulk_files):
                    status_text.text(f"Processing {i+1}/{len(bulk_files)}: {f.name}…")
                    try:
                        buf, bid = process_file(f, marketplace_type)
                        out_name = get_output_filename(f.name, bid)
                        results.append((out_name, buf))
                    except Exception as e:
                        errors.append(f"`{f.name}`: {e}")
                    progress.progress((i + 1) / len(bulk_files))

                status_text.empty()
                progress.empty()

                if results:
                    first_bid = int(results[0][0].split("batch")[1].replace(".xlsx", ""))
                    last_bid  = int(results[-1][0].split("batch")[1].replace(".xlsx", ""))
                    zip_buf = build_zip(results)
                    st.success(f"✅ {len(results)} file(s) processed! BatchIDs {first_bid}–{last_bid}")
                    st.download_button(
                        f"📥 Download ZIP ({len(results)} files)",
                        data=zip_buf,
                        file_name=f"bulk_outputs_batch{first_bid}_to_{last_bid}.zip",
                        mime="application/zip",
                        key="bulk_download"
                    )
                if errors:
                    st.error("Some files failed:")
                    for e in errors:
                        st.markdown(f"- {e}")

# ══════════════════════════════════════════════════════════════
#  SINGLE FILE MODE (original behaviour)
# ══════════════════════════════════════════════════════════════
else:
    st.markdown("---")

    flipkart_catalog_file = None
    flipkart_listing_file = None

    if marketplace_type == "Flipkart":
        st.info("📂 Upload one or both Flipkart files. At least one is required.")
        col_cat, col_lst = st.columns(2)
        with col_cat:
            flipkart_catalog_file = st.file_uploader("Catalog File", type=["xlsx", "xls", "xlsm"], key="fk_catalog")
        with col_lst:
            flipkart_listing_file = st.file_uploader("Listing File", type=["xlsx", "xls", "xlsm"], key="fk_listing")
        input_file = flipkart_catalog_file or flipkart_listing_file
    else:
        input_file = st.file_uploader("Upload Input Excel File", type=["xlsx", "xls", "xlsm"])

    selected_variant_col = "(none)"
    selected_product_col = "(none)"

    if input_file:
        selected_sheet = None
        if marketplace_type == "General":
            try:
                xl     = pd.ExcelFile(input_file)
                sheets = xl.sheet_names
                selected_sheet = st.selectbox("Select sheet", sheets)
            except Exception as e:
                st.error(f"Failed to read sheets from uploaded file: {e}")
                selected_sheet = None

        try:
            if marketplace_type == "Flipkart":
                src_df = merge_flipkart_files(
                    catalog_file=flipkart_catalog_file,
                    listing_file=flipkart_listing_file
                )
            else:
                src_df = read_input_to_df(
                    input_file, marketplace_type,
                    header_row=general_header_row,
                    data_row=general_data_row,
                    sheet_name=selected_sheet
                )
        except Exception as e:
            st.error(f"Failed to parse uploaded file: {e}")
            src_df = None

        if src_df is not None:
            if marketplace_type == "General":
                st.markdown("**Sample data (first 3 rows)**")
                st.dataframe(src_df.head(3))
                cols = ["(none)"] + [str(c) for c in src_df.columns]
                col1, col2 = st.columns(2)
                with col1:
                    selected_variant_col = st.selectbox("Style Code → productId (leave '(none)' to skip)", options=cols, index=0)
                with col2:
                    selected_product_col = st.selectbox("Seller SKU → variantId (leave '(none)' to skip)", options=cols, index=0)
            else:
                if marketplace_type == "Amazon":
                    filtered = src_df.attrs.get("filtered_parent_rows", 0)
                    if filtered:
                        st.info(f"ℹ️ {filtered} Parent row(s) removed (Parentage Level = 'Parent')")
                st.subheader("Preview (first 5 rows)")
                try:
                    st.dataframe(src_df.head(5))
                except Exception as e:
                    st.warning(f"Could not render preview: {e}")
                    st.write(src_df.head(5).to_string())

        st.markdown("---")

        if marketplace_type == "General":
            if st.button("Generate Output"):
                with st.spinner("Processing…"):
                    try:
                        result, assigned_batch_id = process_file(
                            input_file, marketplace_type,
                            selected_variant_col=selected_variant_col,
                            selected_product_col=selected_product_col,
                            general_header_row=general_header_row,
                            general_data_row=general_data_row,
                            general_sheet_name=selected_sheet,
                        )
                        if result:
                            out_name = get_output_filename(input_file.name, assigned_batch_id)
                            st.success(f"✅ Output Generated! — BatchID assigned: **{assigned_batch_id}**")
                            st.download_button(
                                "📥 Download Output",
                                data=result,
                                file_name=out_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_button"
                            )
                    except Exception as e:
                        st.error(f"Processing failed: {e}")
        else:
            with st.spinner("Processing…"):
                try:
                    _premerged = None
                    if marketplace_type == "Flipkart":
                        _premerged = merge_flipkart_files(
                            catalog_file=flipkart_catalog_file,
                            listing_file=flipkart_listing_file
                        )
                    result, assigned_batch_id = process_file(
                        input_file, marketplace_type,
                        selected_variant_col=None,
                        selected_product_col=None,
                        general_header_row=general_header_row,
                        general_data_row=general_data_row,
                        general_sheet_name=None,
                        premerged_df=_premerged,
                    )
                    if result:
                        out_name = get_output_filename(input_file.name, assigned_batch_id)
                        st.success(f"✅ Output Generated! — BatchID assigned: **{assigned_batch_id}**")
                        st.download_button(
                            "📥 Download Output",
                            data=result,
                            file_name=out_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_button"
                        )
                except Exception as e:
                    st.error(f"Processing failed: {e}")
    else:
        st.info("Upload a file to get started.")

st.markdown("---")
st.caption("Built for Rubick.ai | By Vishnu Sai")
